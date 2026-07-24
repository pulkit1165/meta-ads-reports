#!/usr/bin/env python3
"""
build_products_report.py — "All Portals · Products Report" as a downloadable Excel.

This is a faithful port of Navdeep's manual products-report script (the one that
fills the "All Portals · DD Mon · Products Report" tab in the Daily Camp Pushed
sheet). Verified 24 Jul 2026 to reproduce that sheet's 10-day columns exactly
(attempts + revenue exact, spend within 0.01%).

Why a separate live pull (not our ntn.db):
  Navdeep pulls LIVE at CAMPAIGN level from 6 delivering accounts with a
  hand-built product map. Our ntn.db is ad-level and under-counts campaign
  coverage, so it can't match the sheet. This pull is light — 6 accounts ×
  (campaigns + today insights + 10d insights) ≈ 36 API calls, a few seconds —
  so it is safe to run hourly and will not hang the pipeline the way the heavy
  ad-level-across-75-accounts pull did.

Dates are DYNAMIC:
  * today = current IST date (or --date YYYY-MM-DD)
  * 10d   = the 10 completed days ending yesterday (date-10 .. date-1)
    (matches Navdeep's hardcoded today=23 / 10d=13..22)

Output:
  * <out>.xlsx           — styled workbook matching the sheet's columns
  * <out>.json           — sidecar {generated, grand:{...}} for the dashboard panel

Best-effort: prints a warning and exits 0 on failure so it can never break the
deploy it runs inside.
"""
from __future__ import annotations
import argparse, json, os, sys, time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

IST = timezone(timedelta(hours=5, minutes=30))
G = "https://graph.facebook.com/v19.0"
ATTRIB = json.dumps(["1d_view", "7d_click"])

# The 6 delivering accounts, hardcoded to their portal (verbatim from Navdeep's
# script). These are the accounts that actually deliver for this report; the
# report is deliberately a curated subset, not every configured account.
ACCTS = [
    ('SM Fragrance',     'act_466922745634023',   'SM'),
    ('SM Crystals',      'act_1181596092752041',  'SM'),
    ('NBP Hair/Perfume', 'act_1501832634098072',  'NBP'),
    ('NBP Skin',         'act_1505319823511657',  'NBP'),
    ('SML Skin',         'act_918587349998103',   'SML'),
    ('SML Hair',         'act_1229831035065328',  'SML'),
]

# ── Product map (verbatim from Navdeep's script — this IS the report's definition) ──
REMAP_SM = '120246532363410384'   # force this campaign into 7 Chakra Chain 549
SHIVA = {'120246775180300436', '120246775061870436'}
IMPRESSION = {'120251255369280316'}


def prod_sm(n, cid):
    if cid == REMAP_SM: return '7 Chakra Chain 549'
    l = n.lower()
    if 'ntn1682' in l: return '7 Chakra Chain 549'
    if 'ntn1686' in l or 'rope_chain' in l: return 'Rope Chain 899'
    if 'ntn1577' in l or 'richie_rich_bracelet_499' in l or 'rich_bracelet_499' in l: return 'RR Gold Bracelet 499'
    if 'ntn1687' in l or ('rice' in l and '399' in l): return 'RR Gold Bracelet 399'
    if 'ntn1679' in l: return '7 Chakra Bracelet 499'
    if 'ntn1290' in l or 'velore' in l: return 'Velore Chain'
    if 'ntn222' in l or 'richie_rich_combo' in l: return 'RR Half n Half (SM Crystal)' if 'crystal' in l.lower() or cid.endswith('581') else 'RR Combo 999'
    if 'ntn1990' in l or 'keychain' in l or 'new_keyring' in l: return 'New Keyring'
    if 'richie_rice_chain' in l or 'rice_chain_549' in l: return 'RR Gold Chain 549'
    if 'ntn1716' in l or 'gold_chips' in l: return 'Chips Bracelet (new)'
    if 'geode' in l: return 'Geode'
    if 'ntn950' in l or 'unstoppable' in l: return 'Unstoppable Bracelet Crystal'
    if '299_collection' in l or 'sale_299' in l: return '299 Collection'
    if 'catalog_chain' in l or 'catlog_chain' in l: return 'Catalog Chain'
    if 'ntn1217' in l or 'selenite' in l: return 'Selenite Plate'
    if 'ntn1523' in l or ('pyrite' in l and ('keyring' in l or 'agate' in l)): return 'Pyrite & Agate Keyring'
    if 'ntn229' in l or '7_chakra_bracelet' in l: return '7 Chakra Bracelet 229'
    if 'catalog_pendant' in l or 'catlog_pendant' in l: return 'Catalog Pendant'
    if 'carousel' in l or 'corousel' in l: return 'Carousel Chain'
    return None


def prod_nbp(n, cid):
    if cid in IMPRESSION: return 'Impression Retarget'
    l = n.lower()
    if 'imp_rtg' in l or 'impression' in l: return 'Impression Retarget'
    if 'ampm' in l or 'ntn237' in l: return 'AM PM'
    if 'sunkissed' in l or 'ntn402' in l: return 'Sun Kissed / Bubble Gum'
    if 'nona_peptide' in l or 'peptide' in l or 'ntn1190' in l or 'ntn1230' in l: return 'Peptide'
    if 'phusphus_combo' in l or 'ntn761' in l or '24/7' in l or 'all_day_phus_phus' in l: return '24/7 Phus Phus Combo'
    if 'ayurvedic_phusphus' in l or ('phusphus' in l and 'combo' not in l): return 'Ayurvedic Phus Phus Rice Hair Mist'
    if 'phus_phus' in l or 'phusphus' in l: return 'Ayurvedic Phus Phus Rice Hair Mist'
    if 'hyaluronic' in l or 'ntn364' in l: return 'Hyaluronic Water Gel'
    if 'time_reversal' in l or 'ntn221' in l: return 'Time Reversal Combo'
    if 'boba' in l or 'ntn1486' in l: return 'Boba Cream'
    if 'goat_milk' in l or 'puffy' in l or 'under_eye' in l or 'undereye' in l: return 'Under Eye Goat Milk / Puffy Eyes'
    if 'cooling' in l: return 'Cooling Fluid'
    if 'browgrow' in l or 'ntn180' in l: return 'Brow Grow'
    if 'berberine' in l or 'ntn294' in l: return 'Berberine'
    if 'oil+shampoo' in l or 'oil_shampoo' in l or 'fermented_rice' in l: return 'Star Hair Combo'
    return None


def prod_sml(n, cid):
    if cid in SHIVA: return 'Lord Shiva'
    l = n.lower()
    if 'pack_of_2_frame' in l or ('pack_of_2' in l and 'frame' in l): return 'Pack of 2 (Frame Combo)'
    if 'ntn822' in l or 'horse_frame' in l or '7_horse_frame' in l or ('horse' in l and 'frame' in l): return 'Horse Frame 599'
    if 'trishul' in l or 'mahadev' in l: return 'Trishul Frame 649'
    if 'ntn932' in l or 'peacock_frame' in l or 'peacock' in l: return 'Pyrite-Lapis Peacock Frame'
    if 'sale299' in l or '299_collection' in l: return '299 Collection'
    if '399_collection' in l: return '399 Collection'
    if 'shiva' in l: return 'Lord Shiva'
    if 'ntn322' in l or 'pitglow' in l or 'pit_glow' in l: return 'Glass Skin PitGlow Deo P2'
    if 'yellow_phus' in l: return 'Yellow Phus Phus'
    if 'ntn761' in l or 'all_day_phus_phus' in l or '24/7' in l: return '24/7 Phus Phus Combo (761)'
    if 'ntn760' in l or 'xtreme_phus' in l: return 'Xtreme Phus Phus (760)'
    return None


def bucket(name, portal, cid):
    return {'SM': prod_sm, 'NBP': prod_nbp, 'SML': prod_sml}[portal](name, cid)


# Fixed display order per portal (verbatim). Products always listed in this order.
ORDERS = {
    'SM': ['7 Chakra Chain 549', 'Rope Chain 899', 'RR Gold Bracelet 499', '7 Chakra Bracelet 499', 'Velore Chain',
           'RR Gold Bracelet 399', 'RR Combo 999', 'New Keyring', 'RR Gold Chain 549', 'Chips Bracelet (new)',
           'Geode', 'Unstoppable Bracelet Crystal', '299 Collection', 'Catalog Chain', 'Selenite Plate',
           'RR Half n Half (SM Crystal)', 'Pyrite & Agate Keyring', '7 Chakra Bracelet 229', 'Catalog Pendant', 'Carousel Chain'],
    'NBP': ['AM PM', 'Sun Kissed / Bubble Gum', 'Peptide', 'Ayurvedic Phus Phus Rice Hair Mist',
            'Hyaluronic Water Gel', 'Time Reversal Combo', 'Boba Cream', '24/7 Phus Phus Combo',
            'Under Eye Goat Milk / Puffy Eyes', 'Cooling Fluid', 'Brow Grow', 'Berberine', 'Star Hair Combo', 'Impression Retarget'],
    'SML': ['Horse Frame 599', 'Trishul Frame 649', 'Pyrite-Lapis Peacock Frame', '299 Collection', '399 Collection',
            'Lord Shiva', 'Glass Skin PitGlow Deo P2', 'Yellow Phus Phus', '24/7 Phus Phus Combo (761)', 'Xtreme Phus Phus (760)', 'Pack of 2 (Frame Combo)'],
}

SKU = {
    '7 Chakra Chain 549': 'NTN1682', 'Rope Chain 899': 'NTN1686', 'RR Gold Bracelet 499': 'NTN1577',
    '7 Chakra Bracelet 499': 'NTN1679', 'Velore Chain': 'NTN1290', 'RR Gold Bracelet 399': 'NTN1687',
    'RR Combo 999': 'NTN222 (Frag)', 'New Keyring': 'NTN1990', 'RR Gold Chain 549': 'Wanda RR Chain',
    'Chips Bracelet (new)': 'NTN1716', 'Geode': 'Geode', 'Unstoppable Bracelet Crystal': 'NTN950',
    '299 Collection': '299 Collection', 'Catalog Chain': 'Catalog', 'Selenite Plate': 'NTN1217',
    'RR Half n Half (SM Crystal)': 'NTN222 (Cr)', 'Pyrite & Agate Keyring': 'NTN1523',
    '7 Chakra Bracelet 229': 'NTN229', 'Catalog Pendant': 'Catalog Pendant', 'Carousel Chain': 'Carousel',
    'AM PM': 'NTN237', 'Sun Kissed / Bubble Gum': 'NTN402', 'Peptide': 'NTN1190',
    'Ayurvedic Phus Phus Rice Hair Mist': 'PhusPhus Mist', 'Hyaluronic Water Gel': 'NTN364',
    'Time Reversal Combo': 'NTN221', 'Boba Cream': 'NTN1486', '24/7 Phus Phus Combo': 'NTN761',
    'Under Eye Goat Milk / Puffy Eyes': 'Goat Milk', 'Cooling Fluid': 'Cooling Cream',
    'Brow Grow': 'NTN180', 'Berberine': 'NTN294', 'Star Hair Combo': 'Fermented Rice + Oil', 'Impression Retarget': 'Impression RTG',
    'Horse Frame 599': 'NTN822', 'Trishul Frame 649': 'Trishul/Mahadev', 'Pyrite-Lapis Peacock Frame': 'NTN932',
    '399 Collection': 'Neutra 399', 'Lord Shiva': 'Miniature Shiv',
    'Glass Skin PitGlow Deo P2': 'NTN322', 'Yellow Phus Phus': 'Yellow',
    '24/7 Phus Phus Combo (761)': 'NTN761', 'Xtreme Phus Phus (760)': 'NTN760',
    'Pack of 2 (Frame Combo)': 'Pack of 2 Frame',
}

COLS = ['Portal', 'Product', 'SKU', 'Active Camps', 'Active Budget', 'Active Spend',
        'Paused Camps', 'Paused Budget', 'Paused Spend', 'Total Spend', 'Total Rev',
        'Today ROAS', '10d ROAS', '10d Attempts', '10d Spend', '10d Rev']


def paginate(requests, url, params):
    out = []
    while url:
        try:
            r = requests.get(url, params=params, timeout=60).json()
        except Exception:
            time.sleep(5); continue
        if 'error' in r:
            if r['error'].get('code') in (1, 2, 17, 32, 613, 80004):
                time.sleep(10); continue
            break
        out += r.get('data', [])
        nxt = r.get('paging', {}).get('next'); url = nxt if nxt else None; params = None
    return out


def pull_insights(requests, token, since, until):
    out = {}
    for _, acct, _ in ACCTS:
        for i in paginate(requests, f'{G}/{acct}/insights', {
                'access_token': token, 'level': 'campaign',
                'time_range': json.dumps({'since': since, 'until': until}),
                'fields': 'campaign_id,spend,action_values',
                'action_attribution_windows': ATTRIB, 'limit': 500}):
            cid = i['campaign_id']; sp = float(i.get('spend', 0)); rv = 0
            for av in i.get('action_values', []) or []:
                if av.get('action_type') == 'omni_purchase':
                    try: rv = float(av.get('value', 0) or 0)
                    except Exception: pass
                    break
            out[cid] = {'sp': sp, 'rv': rv}
    return out


def compute(requests, token, day: date):
    today_s = day.isoformat()
    d10_s = (day - timedelta(days=10)).isoformat()
    d10_e = (day - timedelta(days=1)).isoformat()

    cid_info = {}
    for _, acct, portal in ACCTS:
        for c in paginate(requests, f'{G}/{acct}/campaigns', {
                'access_token': token, 'fields': 'id,name,effective_status,daily_budget', 'limit': 500}):
            cid_info[c['id']] = {'name': c['name'], 'portal': portal,
                                 'st': c.get('effective_status'),
                                 'db': int(c.get('daily_budget', 0) or 0) / 100}

    today = pull_insights(requests, token, today_s, today_s)
    d10 = pull_insights(requests, token, d10_s, d10_e)

    pp = {p: {} for p in ('SM', 'NBP', 'SML')}
    for cid, c in cid_info.items():
        p = bucket(c['name'], c['portal'], cid)
        if p is None: continue
        portal = c['portal']; is_active = c['st'] == 'ACTIVE'
        t = today.get(cid, {'sp': 0, 'rv': 0}); d = d10.get(cid, {'sp': 0, 'rv': 0})
        if not is_active and t['sp'] <= 0 and d['sp'] <= 0: continue
        x = pp[portal].setdefault(p, {'a_n': 0, 'a_db': 0, 'a_sp': 0, 'a_rv': 0,
                                      'p_n': 0, 'p_db': 0, 'p_sp': 0, 'p_rv': 0,
                                      'd10_n': 0, 'd10_sp': 0, 'd10_rv': 0})
        if is_active:
            x['a_n'] += 1; x['a_db'] += c['db']; x['a_sp'] += t['sp']; x['a_rv'] += t['rv']
        elif t['sp'] > 0:
            x['p_n'] += 1; x['p_db'] += c['db']; x['p_sp'] += t['sp']; x['p_rv'] += t['rv']
        if d['sp'] > 0:
            x['d10_n'] += 1; x['d10_sp'] += d['sp']; x['d10_rv'] += d['rv']
    return pp, (today_s, d10_s, d10_e)


def _row(portal, name, x):
    ts = x['a_sp'] + x['p_sp']; tr = x['a_rv'] + x['p_rv']
    roas = tr / ts if ts else 0
    d10r = x['d10_rv'] / x['d10_sp'] if x['d10_sp'] else 0
    return [portal, name, SKU.get(name, ''), x['a_n'], round(x['a_db']), round(x['a_sp']),
            x['p_n'], round(x['p_db']), round(x['p_sp']), round(ts), round(tr),
            round(roas, 2), round(d10r, 2), x['d10_n'], round(x['d10_sp']), round(x['d10_rv'])]


def build_rows(pp):
    """Returns (rows, meta) where rows include per-portal blocks, subtotals, grand."""
    ZERO = {'a_n': 0, 'a_db': 0, 'a_sp': 0, 'a_rv': 0, 'p_n': 0, 'p_db': 0,
            'p_sp': 0, 'p_rv': 0, 'd10_n': 0, 'd10_sp': 0, 'd10_rv': 0}
    rows = []; subtotal_idx = []; grand_idx = None
    gt = dict(ZERO); gt_ts = gt_tr = 0
    for portal in ('SM', 'NBP', 'SML'):
        st = dict(ZERO); st_ts = st_tr = 0
        for name in ORDERS[portal]:
            x = pp[portal].get(name, ZERO)
            rows.append(_row(portal, name, x))
            for k in ZERO: st[k] += x[k]
            st_ts += x['a_sp'] + x['p_sp']; st_tr += x['a_rv'] + x['p_rv']
        sub = _row(f'{portal} · SUBTOTAL', '', st)
        sub[1] = ''; sub[2] = ''
        subtotal_idx.append(len(rows)); rows.append(sub)
        rows.append(None)  # spacer
        for k in ZERO: gt[k] += st[k]
        gt_ts += st_ts; gt_tr += st_tr
    grand = _row('GRAND TOTAL', '', gt)
    grand[1] = ''; grand[2] = ''
    grand_idx = len(rows); rows.append(grand)
    meta = {'grand': {'active_camps': gt['a_n'], 'active_budget': round(gt['a_db']),
                      'today_spend': round(gt_ts), 'today_rev': round(gt_tr),
                      'today_roas': round(gt_tr / gt_ts, 2) if gt_ts else 0,
                      'd10_spend': round(gt['d10_sp']), 'd10_rev': round(gt['d10_rv']),
                      'd10_roas': round(gt['d10_rv'] / gt['d10_sp'], 2) if gt['d10_sp'] else 0}}
    return rows, subtotal_idx, grand_idx, meta


def write_xlsx(out_path: Path, rows, subtotal_idx, grand_idx, stamp):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    wb = Workbook(); ws = wb.active; ws.title = 'Products Report'
    title = f'Report generated: {stamp} · Meta Ads · Attribution 1d view + 7d click'
    ws.append([title]); ws.append([]); ws.append(COLS)

    HDR = PatternFill('solid', fgColor='1F4E78')
    HDRF = Font(bold=True, color='FFFFFF', size=10)
    TITLEF = Font(bold=True, color='1F4E78', size=12)
    SUB = PatternFill('solid', fgColor='DCE6F1'); SUBF = Font(bold=True)
    GRAND = PatternFill('solid', fgColor='FFE699'); GRANDF = Font(bold=True, size=11)
    thin = Side(style='thin', color='D9D9D9'); border = Border(bottom=thin)

    ws['A1'].font = TITLEF
    for c in ws[3]:
        c.fill = HDR; c.font = HDRF
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[3].height = 30

    r = 4
    excel_subtotal_rows, excel_grand_row = [], None
    for i, row in enumerate(rows):
        if row is None:
            r += 1; continue
        ws.append(row)
        is_sub = i in subtotal_idx
        is_grand = (i == grand_idx)
        if is_sub:
            excel_subtotal_rows.append(r)
            for c in ws[r]: c.fill = SUB; c.font = SUBF
        elif is_grand:
            excel_grand_row = r
            for c in ws[r]: c.fill = GRAND; c.font = GRANDF
        else:
            for c in ws[r]: c.border = border
        r += 1
    last = r - 1

    # number formats
    for col in ('E', 'F', 'H', 'I', 'J', 'K', 'O', 'P'):
        for cell in ws[col][3:]:
            cell.number_format = '#,##0'
    for col in ('L', 'M'):
        for cell in ws[col][3:]:
            cell.number_format = '0.00'

    # ROAS color scales (Today ROAS = L, 10d ROAS = M)
    scale = lambda: ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                                   mid_type='num', mid_value=1.3, mid_color='FFEB84',
                                   end_type='num', end_value=2.5, end_color='63BE7B')
    if last >= 4:
        ws.conditional_formatting.add(f'L4:L{last}', scale())
        ws.conditional_formatting.add(f'M4:M{last}', scale())

    widths = [16, 34, 16, 11, 12, 12, 11, 12, 12, 12, 12, 10, 9, 11, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:P{last}'
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--out', default='roas-live/reports/products-report.xlsx')
    ap.add_argument('--date', default=None, help='YYYY-MM-DD (default: today IST)')
    ap.add_argument('--token', default=None, help='Meta token (default: env META_ACCESS_TOKEN)')
    args = ap.parse_args()

    token = args.token or os.environ.get('META_ACCESS_TOKEN')
    if not token:
        print('  products-report: no META_ACCESS_TOKEN — skipping (non-fatal)'); return 0
    try:
        import requests
    except ImportError:
        print('  products-report: requests not installed — skipping'); return 0

    now = datetime.now(IST)
    day = datetime.strptime(args.date, '%Y-%m-%d').date() if args.date else now.date()
    stamp = f'{day:%d %b %Y} · {now:%I:%M %p} IST'

    try:
        pp, (today_s, d10_s, d10_e) = compute(requests, token, day)
        rows, sub_idx, grand_idx, meta = build_rows(pp)
        out = Path(args.out)
        write_xlsx(out, rows, sub_idx, grand_idx, stamp)
        g = meta['grand']
        side = {'generated': now.isoformat(), 'stamp': stamp,
                'today': today_s, 'window_10d': f'{d10_s}..{d10_e}',
                'file': out.name, **meta}
        out.with_suffix('.json').write_text(json.dumps(side, indent=1))
        print(f'  products-report: wrote {out} — GRAND today {g["today_roas"]} '
              f'(Rs{g["today_spend"]:,}/Rs{g["today_rev"]:,}) · 10d {g["d10_roas"]} '
              f'· active budget Rs{g["active_budget"]:,}')
    except Exception as e:
        import traceback
        print(f'  products-report: FAILED (non-fatal) — {e}'); traceback.print_exc()
    return 0


if __name__ == '__main__':
    sys.exit(main())
