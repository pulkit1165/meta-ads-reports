#!/usr/bin/env python3
"""
xlsx_out.py — shared Excel styling for the automated reports.

Deliberately mirrors the look of the EC2 `simple_camp_db.py` exports the team
already opens daily: navy header band, a title line above it, frozen panes,
autofilter, zebra striping, and ₹/%/ROAS number formats. Verdict and delta
columns are colour-coded so the action is visible without reading numbers.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = PatternFill('solid', fgColor='1F4E78')
TITLE_FILL = PatternFill('solid', fgColor='12355B')
BAND = PatternFill('solid', fgColor='7F6000')          # section banner (amber)
ZEBRA = PatternFill('solid', fgColor='F4F7FB')

FILL_PAUSE = PatternFill('solid', fgColor='FDE9E7')
FILL_REVIEW = PatternFill('solid', fgColor='FFF3CD')
FILL_WATCH = PatternFill('solid', fgColor='FFF9E6')
FILL_SCALE = PatternFill('solid', fgColor='D7EAD9')
FILL_GAP = PatternFill('solid', fgColor='EFEFEF')

VERDICT_FILL = {
    'PAUSE': FILL_PAUSE, 'PAUSE (not whitelisted)': FILL_PAUSE,
    'REVIEW': FILL_REVIEW, 'WATCH': FILL_WATCH, 'SCALE': FILL_SCALE,
}

HEAD_FONT = Font(bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(bold=True, color='FFFFFF', size=11)
BAND_FONT = Font(bold=True, color='FFFFFF', size=10)

_thin = Side(style='thin', color='D0D7E5')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')


def _autosize(ws, header, rows, cap=46):
    for i, h in enumerate(header, start=1):
        width = len(str(h))
        for r in rows:
            if i - 1 < len(r):
                width = max(width, len(str(r[i - 1] if r[i - 1] is not None else '')))
        ws.column_dimensions[get_column_letter(i)].width = min(cap, max(9, width + 2))


def write_table(ws, title: str, header: list, rows: list, *, numfmt: dict | None = None,
                center_cols: set | None = None, verdict_col: int | None = None,
                start: int = 1, band: str | None = None) -> int:
    """Write a title + header + rows block. Returns the next free row."""
    ncol = len(header)
    r = start
    if title:
        ws.cell(row=r, column=1, value=title)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        c = ws.cell(row=r, column=1)
        c.fill = TITLE_FILL
        c.font = TITLE_FONT
        c.alignment = LEFT
        ws.row_dimensions[r].height = 22
        r += 2
    if band:
        ws.cell(row=r, column=1, value=band)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        c = ws.cell(row=r, column=1)
        c.fill = BAND
        c.font = BAND_FONT
        r += 1

    head_row = r
    for i, h in enumerate(header, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.fill = NAVY
        c.font = HEAD_FONT
        c.alignment = CENTER
        c.border = BORDER
    r += 1

    numfmt = numfmt or {}
    center_cols = center_cols or set()
    for n, row in enumerate(rows):
        vfill = VERDICT_FILL.get(str(row[verdict_col])) if verdict_col is not None else None
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER
            c.alignment = CENTER if i in center_cols else LEFT
            if i in numfmt:
                c.number_format = numfmt[i]
            if vfill is not None:
                c.fill = vfill
            elif v == 'no snapshot':
                c.fill = FILL_GAP
            elif n % 2:
                c.fill = ZEBRA
        r += 1

    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)
    if rows:
        ws.auto_filter.ref = (f'A{head_row}:{get_column_letter(ncol)}{head_row + len(rows)}')
    _autosize(ws, header, rows)
    return r + 1


def new_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb
