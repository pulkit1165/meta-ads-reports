#!/usr/bin/env python3
"""Fill Creative / Product name / SKU name for an xlsx whose first column is
Campaign ID. Builds a campaign_id -> (creative, product, sku) map from the DB:
  creative = dominant creative type (by spend) across the campaign's ad names
             [paras/ugc/partnership/static/motion], from meta_ads_daily
  product  = dominant product (by spend), same pipeline as the report
  sku      = NTN code in the campaign's names, else the product's crosswalk SKU
Campaigns with no ad rows fall back to product/SKU from the campaign name (meta_
campaigns); creative is left blank (no ad name to read). IDs in neither table
stay blank.

Usage: python3 scripts/v2/fill_campaign_ids_xlsx.py "<input.xlsx>" ["<out.xlsx>"]
"""
import sqlite3
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

_REPO = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(_REPO / "scripts"))
sys.path.insert(0, str(_REPO / "scripts" / "v2"))

from classify_ads import extract_ntn_code  # noqa: E402
from map_products_xlsx import (load_ntn_map, load_crosswalk, classify_row,  # noqa: E402
                               creative_type_from_adname, DB)


def _dominant(d):
    return max(d, key=d.get) if d else ""


def build_campaign_map(db, ntn_map, cross):
    con = sqlite3.connect(db)
    con.row_factory = sqlite3.Row
    agg = defaultdict(lambda: {"ct": defaultdict(float), "pr": defaultdict(float),
                               "names": set()})
    for r in con.execute("SELECT campaign_id, campaign_name, ad_name, spend "
                          "FROM meta_ads_daily"):
        cid = str(r["campaign_id"]); a = agg[cid]
        sp = r["spend"] or 0
        a["ct"][creative_type_from_adname(r["ad_name"] or "")] += sp
        a["pr"][classify_row(r["ad_name"] or "", r["campaign_name"] or "", ntn_map)] += sp
        a["names"].add(f"{r['ad_name'] or ''} {r['campaign_name'] or ''}")
    camp_name = {str(r["campaign_id"]): r["name"]
                 for r in con.execute("SELECT campaign_id, name FROM meta_campaigns")}
    con.close()

    def sku_of(product, names):
        for nm in names:
            code = extract_ntn_code(nm)
            if code and code in ntn_map:
                return code
        return cross.get(product, "")

    out = {}
    for cid, a in agg.items():            # campaigns with ad data
        product = _dominant(a["pr"])
        out[cid] = {"creative": _dominant(a["ct"]), "product": product,
                    "sku": sku_of(product, a["names"])}
    for cid, name in camp_name.items():   # name-only campaigns (no ads/spend)
        if cid in out:
            continue
        product = classify_row("", name or "", ntn_map)
        out[cid] = {"creative": "", "product": product,
                    "sku": sku_of(product, {name or ""})}
    return out


def main():
    if len(sys.argv) < 2:
        sys.exit("usage: fill_campaign_ids_xlsx.py <input.xlsx> [output.xlsx]")
    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else \
        str(Path(inp).with_name(Path(inp).stem + " - filled.xlsx"))

    ntn_map = load_ntn_map(DB)
    cmap = build_campaign_map(DB, ntn_map, load_crosswalk())
    print(f"campaign map: {len(cmap)} campaigns")

    wb = openpyxl.load_workbook(inp)
    ws = wb.active
    header = [str(c.value or "").strip().lower() for c in ws[1]]

    def col(name):
        return header.index(name) + 1 if name in header else None
    c_id = col("campaign id") or 1
    c_cr = col("creative") or 2
    c_pr = col("product name") or 3
    c_sk = col("sku name") or 4

    filled = blank = 0
    for r in range(2, ws.max_row + 1):
        cid = str(ws.cell(row=r, column=c_id).value or "").replace(".0", "").strip()
        if not cid:
            continue
        m = cmap.get(cid)
        if not m:
            blank += 1
            continue
        ws.cell(row=r, column=c_cr, value=m["creative"])
        ws.cell(row=r, column=c_pr, value=m["product"])
        ws.cell(row=r, column=c_sk, value=m["sku"])
        filled += 1

    wb.save(out)
    print(f"filled {filled} rows · {blank} left blank (campaign id not in DB) "
          f"-> {out}")


if __name__ == "__main__":
    main()
