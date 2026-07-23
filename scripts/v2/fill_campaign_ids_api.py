#!/usr/bin/env python3
"""Backfill the still-blank rows of an id_s-style xlsx (Campaign ID | Creative |
Product name | SKU name) from the Meta Graph API — WITHOUT touching rows that
already have a product. For each blank campaign id we fetch its name + ad names
in one call (field expansion) and derive:
  creative = dominant creative type across ad names (paras/ugc/partnership/...)
  product  = dominant product (same classifier as the report)
  sku      = NTN code in the names, else the product's crosswalk SKU

Usage: python3 scripts/v2/fill_campaign_ids_api.py "<filled.xlsx>"
"""
import os
import sys
import time
import json
import urllib.parse
import urllib.request
from collections import Counter
from pathlib import Path

import openpyxl

_REPO = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(_REPO / "scripts"))
sys.path.insert(0, str(_REPO / "scripts" / "v2"))

from classify_ads import extract_ntn_code  # noqa: E402
from map_products_xlsx import (load_ntn_map, load_crosswalk, classify_row,  # noqa: E402
                               creative_type_from_adname, DB)

GRAPH = "https://graph.facebook.com/v21.0"


def token():
    env = Path(os.path.expanduser("~/.openclaw/workspace/.env")).read_text()
    for line in env.splitlines():
        if line.startswith("META_ACCESS_TOKEN="):
            return line.split("=", 1)[1].strip()
    sys.exit("no META_ACCESS_TOKEN in .env")


def fetch_campaign(cid, tok):
    """Return (name, [ad_names]) or (None, None) on error/no-access."""
    q = urllib.parse.urlencode({"fields": "name,ads.limit(150){name}",
                                "access_token": tok})
    try:
        d = json.load(urllib.request.urlopen(f"{GRAPH}/{cid}?{q}", timeout=30))
    except Exception:
        return None, None
    if "error" in d:
        return None, None
    ads = [a.get("name", "") for a in (d.get("ads", {}).get("data") or [])]
    return d.get("name"), ads


def main():
    if len(sys.argv) < 2:
        sys.exit("usage: fill_campaign_ids_api.py <filled.xlsx>")
    path = sys.argv[1]
    tok = token()
    ntn_map = load_ntn_map(DB)
    cross = load_crosswalk()

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header = [str(c.value or "").strip().lower() for c in ws[1]]

    def col(name, default):
        return header.index(name) + 1 if name in header else default
    c_id, c_cr = col("campaign id", 1), col("creative", 2)
    c_pr, c_sk = col("product name", 3), col("sku name", 4)

    # distinct campaign ids whose product cell is still blank
    blanks = {}
    for r in range(2, ws.max_row + 1):
        cid = str(ws.cell(row=r, column=c_id).value or "").replace(".0", "").strip()
        prod = ws.cell(row=r, column=c_pr).value
        if cid and not prod:
            blanks.setdefault(cid, []).append(r)
    print(f"{len(blanks)} distinct campaign ids to backfill "
          f"({sum(len(v) for v in blanks.values())} rows)")

    resolved, no_access = {}, 0
    for i, cid in enumerate(blanks, 1):
        name, ads = fetch_campaign(cid, tok)
        if name is None:
            no_access += 1
            continue
        cts = Counter(creative_type_from_adname(a) for a in ads if a)
        prods = Counter(classify_row(a, name, ntn_map) for a in ads if a) or \
            Counter([classify_row("", name, ntn_map)])
        product = prods.most_common(1)[0][0]
        creative = cts.most_common(1)[0][0] if cts else ""
        names = list(ads) + [name]
        sku = ""
        for nm in names:
            code = extract_ntn_code(nm or "")
            if code and code in ntn_map:
                sku = code
                break
        if not sku:
            sku = cross.get(product, "")
        resolved[cid] = {"creative": creative, "product": product, "sku": sku}
        if i % 25 == 0:
            print(f"  {i}/{len(blanks)} queried · {len(resolved)} resolved")
        time.sleep(0.15)

    filled = 0
    for cid, rows in blanks.items():
        m = resolved.get(cid)
        if not m:
            continue
        for r in rows:
            ws.cell(row=r, column=c_cr, value=m["creative"])
            ws.cell(row=r, column=c_pr, value=m["product"])
            ws.cell(row=r, column=c_sk, value=m["sku"])
            filled += 1

    wb.save(path)
    print(f"\nbackfilled {filled} rows from API "
          f"({len(resolved)} campaigns resolved, {no_access} no-access/error) "
          f"-> {path}")


if __name__ == "__main__":
    main()
