#!/usr/bin/env python3
"""
build_paras_data.py — enrich the Paras video report with preview URLs → JSON.

Reads the xlsx produced by paras_video_report.py (video rows + deployments),
then fetches each video's thumbnail + Facebook permalink via the account-scoped
advideos edge (direct /{video_id} access is denied to this token, but advideos
is granted). Writes a single JSON the dashboard page embeds.

advideos returns videos newest-first, and every report video is from the last
few weeks, so a bounded page walk per account covers them without listing an
account's entire video history.

The mp4 `source` URL is temporary, so we keep the STABLE `permalink_url` (opens
the video on Facebook) plus the thumbnail `picture` as the poster — a preview
that doesn't rot. Re-run to refresh.

Usage:
  META_ACCESS_TOKEN=... python3 scripts/v2/build_paras_data.py \
      --xlsx ~/Downloads/paras_video_report.xlsx --out antariksh/paras_videos.json
"""
from __future__ import annotations

import argparse
import json
import os
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from pathlib import Path

import openpyxl

API = 'https://graph.facebook.com/v19.0'
REPO = Path(__file__).resolve().parent.parent.parent

ACCT_ID = {  # friendly name (as in the xlsx Account col) -> act_ id
    'SM_FRAGRANCE_01': 'act_466922745634023', 'SM_CRYSTALS': 'act_1181596092752041',
    'SM_SKIN': 'act_578075381064759', 'SM_HAIR': 'act_944634709928295',
    'SM_PERFUME': 'act_935485377999005', 'SML_SKIN': 'act_918587349998103',
    'SML_HAIR': 'act_1229831035065328', 'NBP_SKIN': 'act_1505319823511657',
    'NBP_HAIR_PERFUME': 'act_1501832634098072',
}


def _get(url, retries=6):
    delay = 15
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(url, timeout=90) as r:
                return json.load(r)
        except urllib.error.HTTPError as e:
            if e.code == 403 or attempt == retries - 1:
                raise
        except (urllib.error.URLError, TimeoutError):
            if attempt == retries - 1:
                raise
        time.sleep(delay)
        delay = min(delay * 2, 240)


def preview_map(aid: str, token: str, since: str, until: str) -> dict:
    """{video_id: {permalink, thumb}} from the account's DELIVERING ads.

    The report's video_ids are Page-owned videos, invisible to /advideos, so we
    go through the ad creatives instead: delivering ad_ids from async insights,
    then batch their creatives for thumbnail_url + the story's page id. The watch
    link is facebook.com/{page_id}/videos/{video_id}, which opens the reel.
    """
    from paras_video_report import fetch_daily_ad_insights, batch_ids
    daily = fetch_daily_ad_insights(aid, token, since, until)
    if not daily:
        return {}
    meta = batch_ids(daily.keys(),
                     'creative{video_id,object_story_spec,asset_feed_spec,'
                     'thumbnail_url,effective_object_story_id}', token)
    from paras_video_report import video_id_of
    out = {}
    for _ad_id, m in meta.items():
        if not isinstance(m, dict):
            continue
        cr = m.get('creative') or {}
        vid = video_id_of(cr)
        if not vid:
            continue
        story = cr.get('effective_object_story_id') or ''
        pageid = story.split('_', 1)[0] if '_' in story else ''
        out[vid] = {
            'permalink': f'https://www.facebook.com/{pageid}/videos/{vid}' if pageid else '',
            'thumb': cr.get('thumbnail_url', ''),
        }
    return out


def num(x, d=0.0):
    try:
        return float(x)
    except (TypeError, ValueError):
        return d


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', default=str(Path.home() / 'Downloads' / 'paras_video_report.xlsx'))
    ap.add_argument('--out', default=str(REPO / 'antariksh' / 'paras_videos.json'))
    ap.add_argument('--since', default='2026-06-26')
    ap.add_argument('--until', default='')
    args = ap.parse_args()
    token = os.environ['META_ACCESS_TOKEN']

    wb = openpyxl.load_workbook(args.xlsx)
    # Videos: Category, Name, video_id, Tries, Ads, First launch, Spend, Rev, ROAS, 1d, 3d, 7d
    videos = {}
    for r in wb['Videos'].iter_rows(min_row=3, values_only=True):
        if not r or not r[2] or str(r[2]) == 'video_id':
            continue
        videos[str(r[2])] = {
            'category': r[0], 'name': r[1], 'video_id': str(r[2]),
            'tries': int(num(r[3])), 'ads': int(num(r[4])), 'launch': r[5],
            'spend': round(num(r[6])), 'revenue': round(num(r[7])), 'roas': num(r[8]),
            'roas_1d': num(r[9]), 'roas_3d': num(r[10]), 'roas_7d': num(r[11]),
            'permalink': '', 'thumb': '', 'account': '', 'deployments': [],
        }

    # Deployments: Video, video_id, Campaign, Account, First ad, start, end,
    #              Survived, Status, Spend, Rev, ROAS
    by_acct_videos: dict = defaultdict(set)
    for r in wb['Deployments'].iter_rows(min_row=3, values_only=True):
        if not r or not r[1] or str(r[1]) == 'video_id':
            continue
        vid = str(r[1])
        if vid not in videos:
            continue
        videos[vid]['deployments'].append({
            'campaign': r[2], 'account': r[3], 'survived': (int(num(r[7])) if r[7] not in (None, '') else None),
            'status': r[8], 'spend': round(num(r[9])), 'revenue': round(num(r[10])), 'roas': num(r[11]),
        })
        if not videos[vid]['account']:
            videos[vid]['account'] = r[3]
        if r[3] in ACCT_ID:
            by_acct_videos[r[3]].add(vid)

    # fetch preview urls per account
    print(f'{len(videos)} videos across {len(by_acct_videos)} accounts — fetching preview URLs')
    import sys as _sys
    _sys.path.insert(0, str(Path(__file__).resolve().parent))
    found = 0
    for aname, want in by_acct_videos.items():
        aid = ACCT_ID[aname]
        try:
            m = preview_map(aid, token, args.since, args.until or '2026-07-22')
        except Exception as e:
            print(f'  {aname}: preview fetch failed ({str(e)[:60]})'); continue
        n = 0
        for vid, u in m.items():
            if vid in videos:
                videos[vid].update(u); n += 1
        found += n
        print(f'  {aname:20} {n}/{len(want)} previews', flush=True)
        time.sleep(6)

    # category rollup
    cats = defaultdict(lambda: {'videos': 0, 'tries': 0, 'spend': 0.0, 'revenue': 0.0})
    for v in videos.values():
        c = cats[v['category']]
        c['videos'] += 1; c['tries'] += v['tries']; c['spend'] += v['spend']; c['revenue'] += v['revenue']
    for c in cats.values():
        c['roas'] = round(c['revenue'] / c['spend'], 2) if c['spend'] else 0.0

    tot_spend = sum(v['spend'] for v in videos.values())
    tot_rev = sum(v['revenue'] for v in videos.values())
    payload = {
        'since': args.since, 'until': args.until,
        'totals': {'videos': len(videos), 'spend': round(tot_spend), 'revenue': round(tot_rev),
                   'roas': round(tot_rev / tot_spend, 2) if tot_spend else 0.0,
                   'previews': found},
        'categories': cats,
        'videos': sorted(videos.values(), key=lambda v: -v['spend']),
    }
    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    Path(args.out).write_text(json.dumps(payload, indent=1, default=str))
    print(f'\nwrote {args.out} — {len(videos)} videos, {found} with previews')


if __name__ == '__main__':
    main()
