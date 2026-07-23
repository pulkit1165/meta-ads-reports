#!/usr/bin/env python3
"""Add Product + SKU columns to a Meta 'Raw Data Report' xlsx export.

Reuses the creative-elimination classifier pipeline so the Product/SKU values
match the report exactly:
  product = classify_corrected(campaign|ad) -> jewellery sub-split / Unmapped
            recovery (NTN code -> SKU name, else cleaned keyword slug)
  sku     = NTN code in the row's own name if present, else the product's
            crosswalk SKU (from product_ntn_labels).

Usage:
  python3 scripts/v2/map_products_xlsx.py "<input.xlsx>" ["<output.xlsx>"]
"""
import csv
import glob
import os
import sqlite3
import sys
from pathlib import Path

import openpyxl

_REPO = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(_REPO / "scripts"))
sys.path.insert(0, str(_REPO / "scripts" / "v2"))

from active_budget_by_product import classify_corrected  # noqa: E402
from classify_ads import extract_ntn_code  # noqa: E402
from creative_elimination import jewellery_subproduct, recover_unmapped  # noqa: E402

DB = os.environ.get("NTN_DB", "/tmp/ntn.db")


def load_ntn_map(db):
    con = sqlite3.connect(db)
    m = {c: (p, cat) for c, p, cat in con.execute(
        "SELECT ntn_code, product, category FROM product_ntn_labels")}
    con.close()
    return m


def load_roas_map(db, since="2026-05-26"):
    """ad_id -> blended ROAS over the DB's available window from `since`.
    The DB lags (max ~19 Jun); ads not present return None (token pending)."""
    con = sqlite3.connect(db)
    cutoff = con.execute("SELECT MAX(date) FROM meta_ads_daily").fetchone()[0]
    m = {}
    for aid, rev, sp in con.execute(
            "SELECT ad_id, SUM(revenue), SUM(spend) FROM meta_ads_daily "
            "WHERE date >= ? GROUP BY ad_id", (since,)):
        if sp and sp > 0:
            m[str(aid)] = round((rev or 0) / sp, 2)
    con.close()
    return m, since, cutoff


def load_crosswalk():
    """product -> ntn_code, from the latest report SKU-gaps CSV if present."""
    files = sorted(glob.glob(str(_REPO / "out" / "creative_sku_gaps_*.csv")))
    cross = {}
    if files:
        for r in csv.DictReader(open(files[-1])):
            if r.get("ntn_code"):
                cross[r["product"]] = r["ntn_code"]
    return cross


def creative_type_from_adname(ad_name):
    """Creative type read from the AD NAME only, into the operator's taxonomy:
    partnership / ugc / paras / static / motion (first match wins). Partnership
    and ugc are checked before paras so a paras-edited collab isn't mislabelled;
    'other' when no tag is present."""
    t = (ad_name or "").lower()
    if any(k in t for k in ("partnership", "collab", "creator")):
        return "partnership"
    if "ugc" in t:
        return "ugc"
    if "paras" in t:
        return "paras"
    if any(k in t for k in ("static", "image", "post", "carousel")):
        return "static"
    if any(k in t for k in ("motion", "reel", "video")):
        return "motion"
    return "other"


def classify_row(ad_name, campaign_name, ntn_map):
    """Best product for a row — try campaign, fall back to ad if Unmapped, then
    jewellery sub-split / Unmapped recovery (same logic as the report)."""
    product, _ = classify_corrected(campaign_name or ad_name or "")
    if product == "Unmapped" and ad_name:
        p2, _ = classify_corrected(ad_name)
        if p2 != "Unmapped":
            product = p2
    if product == "Jewellery":
        product = jewellery_subproduct(ad_name, campaign_name)
    elif product == "Unmapped":
        rec = recover_unmapped(ad_name, campaign_name, ntn_map)
        if rec:
            product = rec[0]
    return product


def sku_for(ad_name, campaign_name, product, ntn_map, cross):
    code = extract_ntn_code(ad_name or "") or extract_ntn_code(campaign_name or "")
    if code and code in ntn_map:
        return code
    return cross.get(product, "")


def main():
    if len(sys.argv) < 2:
        sys.exit("usage: map_products_xlsx.py <input.xlsx> [output.xlsx]")
    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else \
        str(Path(inp).with_name(Path(inp).stem + " - mapped.xlsx"))

    ntn_map = load_ntn_map(DB)
    cross = load_crosswalk()
    roas_map, since, cutoff = load_roas_map(DB)
    print(f"NTN labels: {len(ntn_map)} · crosswalk products: {len(cross)} · "
          f"ROAS map: {len(roas_map)} ads ({since}..{cutoff})")

    wb = openpyxl.load_workbook(inp)
    ws = wb.active
    header = [c.value for c in ws[1]]

    def col(name):
        for i, h in enumerate(header):
            if h and str(h).strip().lower() == name:
                return i
        return None
    i_ad, i_cp, i_id = col("ad name"), col("campaign name"), col("ad id")
    i_cid = col("campaign id")
    if i_ad is None and i_cp is None:
        sys.exit(f"no 'Ad name'/'Campaign name' column found in: {header}")

    p_col = ws.max_column + 1
    s_col = ws.max_column + 2
    r_col = ws.max_column + 3
    ws.cell(row=1, column=p_col, value="Product")
    ws.cell(row=1, column=s_col, value="SKU")
    ws.cell(row=1, column=r_col, value=f"ROAS ({since[5:]}–{cutoff[5:]})")

    mapped = blank = roas_hit = 0
    from collections import Counter
    prodc = Counter(); ctc = Counter()
    summary = []   # (campaign_id, product, creative_type, roas) for the slim tab
    for r in range(2, ws.max_row + 1):
        ad = ws.cell(row=r, column=i_ad + 1).value if i_ad is not None else ""
        cp = ws.cell(row=r, column=i_cp + 1).value if i_cp is not None else ""
        ad = str(ad or ""); cp = str(cp or "")
        if not ad and not cp:
            blank += 1
            continue
        product = classify_row(ad, cp, ntn_map)
        sku = sku_for(ad, cp, product, ntn_map, cross)
        ctype = creative_type_from_adname(ad)
        ws.cell(row=r, column=p_col, value=product)
        ws.cell(row=r, column=s_col, value=sku)
        roas = None
        if i_id is not None:
            aid = str(ws.cell(row=r, column=i_id + 1).value or "").replace(".0", "")
            if aid in roas_map:
                roas = roas_map[aid]
                ws.cell(row=r, column=r_col, value=roas)
                roas_hit += 1
        cidv = str(ws.cell(row=r, column=i_cid + 1).value or "").replace(".0", "") \
            if i_cid is not None else ""
        summary.append((cidv, product, ctype, roas))
        mapped += 1
        prodc[product] += 1; ctc[ctype] += 1

    # slim tab: Campaign ID | Product Name | Creative Type | ROAS (one row/ad)
    ws2 = wb.create_sheet("Campaign-Product-Type-ROAS", 0)
    ws2.append(["Campaign ID", "Product Name", "Creative Type",
                f"ROAS ({since[5:]}–{cutoff[5:]})"])
    for cidv, product, ctype, roas in summary:
        ws2.append([cidv, product, ctype, roas])

    wb.save(out)
    print(f"mapped {mapped} rows ({blank} blank) -> {out}")
    print(f"ROAS filled: {roas_hit}/{mapped} (rest blank — token pending / "
          f"after {cutoff})")
    print("creative types (from ad name):", dict(ctc.most_common()))
    print("top products:", dict(prodc.most_common(12)))


if __name__ == "__main__":
    main()
